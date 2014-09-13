'''
This module will read the excel document and create instances of nodes, edges to be passed to NetworkX
'''


import openpyxl as px
import numpy as np
import sys, os

from models import *

#
node_format = {
	'label'		: 1,
	'location'  : 2,
	'type'		: 3,
	'no_1_var'	: 4,
	'position'	: 5
}

# edge_format = {
# 	'from_id'	: 1,
# 	'to_id'		: 2,
# 	'weight'	: 3
# }


def parse_spreadsheet(file_path, row_limit):
	#Open the file for reading
	wb = px.load_workbook(filename = file_path)
	
	#This is the node sheet
	ws1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])

	nodes = []

	for sheet_row in range(2,row_limit):
	# for sheet_row in range(1,10):	


		if ws1.cell(row=sheet_row, column=1).value != None: #If no name, don't parse

			#Collect values
			label 	   = ws1.cell(row=sheet_row, column=node_format['label']).value
			location   = ws1.cell(row=sheet_row, column=node_format['location']).value
			type_no    = ws1.cell(row=sheet_row, column=node_format['type']).value
			size 	   = ws1.cell(row=sheet_row, column=node_format['no_1_var']).value
			position   = ws1.cell(row=sheet_row, column=node_format['position']).value
			
			obj = ClimateChangeActor(label, location)

			obj.add_attributes(	type_no	 = type_no, 
								size 	 = size,
								position = str(position))

			#Conditionally set other values such as Alexa Web Ranking

			nodes.append( obj )


	#Now add edges //Deprecated for now

	# ws2 = wb.get_sheet_by_name(wb.get_sheet_names()[1])
	
	edges = []

	# for sheet_row in range(2,ws2.max_row):
	# 	if ws2.cell(row=sheet_row, column=1) != None: #If no id, don't parse
		
	# 		from_id  = ws2.cell(row=sheet_row, column=edge_format['from_id']).value
	# 		to_id    = ws2.cell(row=sheet_row, column=edge_format['to_id']).value

	# 		edge_obj = ActorRelation(from_id=from_id, to_id=to_id)

	# 		edges.append(edge_obj)

	return nodes, edges



